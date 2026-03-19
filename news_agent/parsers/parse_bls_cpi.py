"""
Parse BLS CPI-U monthly data from the archive ZIP.

Input:  raw/impactful/bls_cpi_archive.zip
Output: normalized/bls_cpi_archive.json

Extracts the "All items" CPI-U index for each month from the 12 monthly
cpi-u-YYYYMM.xlsx files inside the ZIP.  The most recent (latest) month
in each file is used.  Also computes month-over-month and year-over-year
percentage changes.
"""

from __future__ import annotations

import io
import re
import zipfile
from typing import Dict, List, Optional

import openpyxl

from .common import NormalizedRecord, raw_path, write_normalized

_SOURCE_ID = "bls_cpi_archive"
_RAW_FILE = raw_path("impactful", "bls_cpi_archive.zip")
_OUT_FILE = "bls_cpi_archive.json"

# Pattern for CPI-U monthly files (not chained-CPI "c-cpi-u")
_CPIU_RE = re.compile(r"^cpi-u-(\d{6})\.xlsx$")

# Month column headers look like "Jan.\n2022" or "May\n2021"
_MONTH_RE = re.compile(r"^(\w+)\.?\n(\d{4})$")
_MONTH_MAP = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def _find_latest_all_items(wb: openpyxl.Workbook) -> Optional[Dict]:
    """Find the latest-month 'All items' CPI index value in a CPI-U workbook."""
    ws = wb.active
    if ws is None:
        return None

    # Row 5 has month headers; row 8 has "All items" values (observed from data)
    # But to be robust, find the "All items" row and the month header row.
    header_row_idx = None
    all_items_row_idx = None

    for r in range(1, min(12, ws.max_row + 1)):
        for c in range(1, min(5, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if isinstance(val, str) and _MONTH_RE.match(val.strip()):
                header_row_idx = r
                break
        if header_row_idx:
            break

    for r in range(1, min(20, ws.max_row + 1)):
        val = ws.cell(r, 2).value  # column B
        if isinstance(val, str) and val.strip().lower() == "all items":
            all_items_row_idx = r
            break

    if not header_row_idx or not all_items_row_idx:
        return None

    # Walk columns to find the last column with a parseable month and a value
    best = None
    for c in range(3, ws.max_column + 1):
        hdr = ws.cell(header_row_idx, c).value
        val = ws.cell(all_items_row_idx, c).value
        if not hdr or not isinstance(val, (int, float)):
            continue
        hdr_str = str(hdr).strip()
        m = _MONTH_RE.match(hdr_str)
        if not m:
            continue
        month_abbr, year = m.group(1), m.group(2)
        mm = _MONTH_MAP.get(month_abbr[:3])
        if not mm:
            continue
        date_key = f"{year}-{mm}"
        # Keep the latest date
        if best is None or date_key > best["date"]:
            best = {"date": date_key, "cpi_index": round(float(val), 3)}

    return best


def parse() -> List[NormalizedRecord]:
    """Return normalised monthly CPI-U records with MoM/YoY changes."""
    zf = zipfile.ZipFile(str(_RAW_FILE))

    # Collect per-file latest "All items" index
    raw_points: Dict[str, float] = {}
    for name in sorted(zf.namelist()):
        if not _CPIU_RE.match(name):
            continue
        data = zf.read(name)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        result = _find_latest_all_items(wb)
        wb.close()
        if result:
            raw_points[result["date"]] = result["cpi_index"]

    zf.close()

    # Build sorted records with MoM and YoY
    sorted_dates = sorted(raw_points.keys())
    records: List[NormalizedRecord] = []
    for date_key in sorted_dates:
        cpi = raw_points[date_key]
        year, month = date_key.split("-")

        # MoM: previous month
        prev_month = int(month) - 1
        prev_year = int(year)
        if prev_month == 0:
            prev_month = 12
            prev_year -= 1
        prev_key = f"{prev_year:04d}-{prev_month:02d}"
        mom_pct = None
        if prev_key in raw_points and raw_points[prev_key] > 0:
            mom_pct = round((cpi - raw_points[prev_key]) / raw_points[prev_key] * 100, 2)

        # YoY: same month last year
        yoy_key = f"{int(year) - 1:04d}-{month}"
        yoy_pct = None
        if yoy_key in raw_points and raw_points[yoy_key] > 0:
            yoy_pct = round((cpi - raw_points[yoy_key]) / raw_points[yoy_key] * 100, 2)

        records.append({
            "date": date_key,
            "cpi_index": cpi,
            "mom_pct": mom_pct,
            "yoy_pct": yoy_pct,
        })

    return records


def run() -> None:
    records = parse()
    out = write_normalized(_SOURCE_ID, records, _OUT_FILE)
    print(f"[{_SOURCE_ID}] {len(records)} monthly CPI points → {out}")


if __name__ == "__main__":
    run()
